import os
import re
import csv
from typing import List, Optional

from qgis.PyQt.QtCore import QVariant, QTimer
from qgis.PyQt.QtGui import QStandardItemModel, QStandardItem
from qgis.PyQt.QtWidgets import (
    QPushButton, QTableView, QProgressBar,
    QVBoxLayout, QHBoxLayout, QLabel,
    QCheckBox, QComboBox, QRadioButton, QFrame,
)
from qgis.gui import QgsFileWidget
from qgis.core import (
    QgsVectorLayer, QgsFeature, QgsGeometry, QgsPointXY,
    QgsProject, QgsField,
)

from .base_widget import BaseDialog
from ..constants import GEOCODER_LAYER, DEFAULT_CRS, KOREA_CRS, ERROR_MESSAGES
from ..core.thread_workers import GeocodingWorker
from ..utils import ConfigManager, ThemeColors, log_info, log_error

# 한국 주소 패턴 (시·도·군·구·읍·면·동·리·로·길 + 경계)
ADDR_PATTERN = re.compile(
    r'(특별시|광역시|특별자치[시도]|도|시|군|구|읍|면|동|리|로|길)(?:\s|$|\d|번|가)'
)

# 좌표 셀 토큰 구분자 (콤마/세미콜론/공백/탭)
COORD_SPLIT_RE = re.compile(r'[\s,;\t]+')

_STATUS_FADE_MS = 3500

# 입력 모드 상수
MODE_ADDRESS = 'address'
MODE_XY = 'xy'        # 단일 셀 "x,y"
MODE_YX = 'yx'        # 단일 셀 "y,x"
MODE_SEP = 'sep'      # X·Y 컬럼 분리


def _status_color(level: str) -> str:
    if level == 'warn':
        return ThemeColors.status_warn()
    if level == 'error':
        return ThemeColors.status_error()
    return ThemeColors.status_info()


def _hline() -> QFrame:
    """수평 구분선 (이전 .ui의 Line 위젯 대체)."""
    line = QFrame()
    line.setFrameShape(QFrame.Shape.HLine)
    line.setFrameShadow(QFrame.Shadow.Sunken)
    return line


def _vtype(name: str):
    """PyQt5/PyQt6 호환: QVariant.String / QVariant.Type.String"""
    value = getattr(QVariant, name, None)
    if value is not None:
        return value
    type_enum = getattr(QVariant, 'Type', None)
    if type_enum is not None:
        return getattr(type_enum, name)
    raise AttributeError(f"QVariant에 {name} 타입을 찾을 수 없습니다.")


class GeocoderWidget(BaseDialog):
    """엑셀/CSV 파일의 주소를 일괄 지오코딩하여 포인트 레이어 생성"""

    def __init__(self, parent=None):
        super().__init__(parent)

        self._build_ui()

        # 상태
        self._raw_rows: List[List[str]] = []
        self._column_count: int = 0
        self._suggested_col: int = 0
        self.worker: Optional[GeocodingWorker] = None
        self._preview_model: Optional[QStandardItemModel] = None

        # 인라인 status 페이드 타이머
        self._status_timer = QTimer(self)
        self._status_timer.setSingleShot(True)
        self._status_timer.timeout.connect(self._clear_status)

        self._populate_crs_combo()
        self._apply_theme_text()
        self._connect_signals()
        self._on_mode_changed()  # 초기 UI 상태 반영
        self._set_running(False)

    # ------------------------------------------------------------------
    # UI 구성
    # ------------------------------------------------------------------
    def _build_ui(self):
        """코드로 UI 구성 (이전 v_world_dockGeocoder_base.ui 대체)."""
        self.setWindowTitle("엑셀 지오코딩")
        self.setMinimumSize(620, 520)
        self.resize(720, 620)

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 16, 20, 16)
        root.setSpacing(10)

        # 헤더 (브이월드 브랜드 스트립)
        root.addWidget(
            self.make_brand_header(
                "엑셀 지오코딩", "엑셀/CSV 주소를 일괄 지오코딩", ":/icon_geocoder"
            )
        )

        # 1. 파일 선택
        self.step1Caption = self.make_section_caption("1. 파일 선택")
        root.addWidget(self.step1Caption)
        self.mQgsFileWidget = QgsFileWidget()
        self.mQgsFileWidget.setFilter("엑셀/CSV (*.xlsx *.xls *.csv)")
        root.addWidget(self.mQgsFileWidget)
        self.fileInfoLabel = QLabel("파일을 선택하면 여기에 행/열 정보가 표시됩니다.")
        self.fileInfoLabel.setWordWrap(True)
        root.addWidget(self.fileInfoLabel)

        # 2. 입력 모드
        self.step2Caption = self.make_section_caption("2. 입력 모드")
        root.addWidget(self.step2Caption)
        mode_row = QHBoxLayout()
        mode_row.setSpacing(12)
        self.modeAddress = QRadioButton("주소")
        self.modeAddress.setChecked(True)
        self.modeCoordXY = QRadioButton("좌표 (x, y)")
        self.modeCoordXY.setToolTip('한 셀에 "경도, 위도" 형태로 저장된 좌표')
        self.modeCoordYX = QRadioButton("좌표 (y, x)")
        self.modeCoordYX.setToolTip('한 셀에 "위도, 경도" 형태로 저장된 좌표')
        self.modeCoordSeparate = QRadioButton("좌표 (X) + 좌표 (Y)")
        self.modeCoordSeparate.setToolTip("X 값과 Y 값이 각자 다른 컬럼에 저장됨")
        for r in (self.modeAddress, self.modeCoordXY, self.modeCoordYX, self.modeCoordSeparate):
            mode_row.addWidget(r)
        mode_row.addStretch(1)
        root.addLayout(mode_row)

        crs_row = QHBoxLayout()
        crs_row.setSpacing(8)
        self.crsLabel = QLabel("좌표계:")
        crs_row.addWidget(self.crsLabel)
        self.crsCombo = QComboBox()
        self.crsCombo.setMinimumWidth(260)
        self.crsCombo.setToolTip(
            "좌표 모드에서 결과 레이어를 생성할 좌표계. "
            "주소 모드에서는 EPSG:4326으로 고정됩니다."
        )
        crs_row.addWidget(self.crsCombo)
        crs_row.addStretch(1)
        root.addLayout(crs_row)

        # 3. 옵션
        self.step3Caption = self.make_section_caption("3. 옵션")
        root.addWidget(self.step3Caption)
        opts = QHBoxLayout()
        opts.setSpacing(12)
        self.headerCheck = QCheckBox("첫 행을 헤더로 사용")
        self.headerCheck.setChecked(True)
        opts.addWidget(self.headerCheck)
        opts.addStretch(1)
        self.columnLabel = QLabel("주소 컬럼:")
        opts.addWidget(self.columnLabel)
        self.columnCombo = QComboBox()
        self.columnCombo.setMinimumWidth(200)
        self.columnCombo.setEnabled(False)
        self.columnCombo.setToolTip("파일을 먼저 선택하세요")
        self.columnCombo.setPlaceholderText("파일을 먼저 선택하세요")
        opts.addWidget(self.columnCombo)
        self.yColumnLabel = QLabel("Y 컬럼:")
        self.yColumnLabel.setVisible(False)
        opts.addWidget(self.yColumnLabel)
        self.yColumnCombo = QComboBox()
        self.yColumnCombo.setMinimumWidth(160)
        self.yColumnCombo.setEnabled(False)
        self.yColumnCombo.setVisible(False)
        self.yColumnCombo.setToolTip("파일을 먼저 선택하세요")
        self.yColumnCombo.setPlaceholderText("파일을 먼저 선택하세요")
        opts.addWidget(self.yColumnCombo)
        root.addLayout(opts)
        self.autoDetectHint = QLabel(
            "파일을 불러오면 가장 가능성 있는 주소 컬럼을 자동 선택합니다."
        )
        self.autoDetectHint.setWordWrap(True)
        root.addWidget(self.autoDetectHint)

        # 4. 미리보기 및 실행
        self.step4Caption = self.make_section_caption("4. 미리보기 및 실행")
        root.addWidget(self.step4Caption)
        self.tableView = QTableView()
        self.tableView.setAlternatingRowColors(True)
        root.addWidget(self.tableView)

        self.geocoderProgressBar = QProgressBar()
        self.geocoderProgressBar.setValue(0)
        root.addWidget(self.geocoderProgressBar)
        self.currentStatusLabel = QLabel("")
        self.currentStatusLabel.setWordWrap(True)
        root.addWidget(self.currentStatusLabel)

        # 푸터
        root.addWidget(_hline())
        footer = QHBoxLayout()
        footer.setSpacing(8)
        self.statusLabel = QLabel("")
        self.statusLabel.setWordWrap(True)
        footer.addWidget(self.statusLabel)
        footer.addStretch(1)
        self.BTNGeoCancel = QPushButton("취소")
        self.BTNGeoCancel.setMinimumWidth(80)
        self.BTNGeoCancel.setVisible(False)
        footer.addWidget(self.BTNGeoCancel)
        self.BTNGeoStart = QPushButton("지오코딩 시작")
        self.BTNGeoStart.setMinimumWidth(120)
        self.BTNGeoStart.setMinimumHeight(32)
        start_font = self.BTNGeoStart.font()
        start_font.setBold(True)
        self.BTNGeoStart.setFont(start_font)
        footer.addWidget(self.BTNGeoStart)
        self.closeBtn = QPushButton("닫기")
        self.closeBtn.setMinimumWidth(80)
        footer.addWidget(self.closeBtn)
        root.addLayout(footer)

    def _apply_theme_text(self):
        """헤더 부제·캡션·힌트 라벨에 테마 대응 색상 적용"""
        muted = ThemeColors.muted()
        if hasattr(self, 'headerSubtitle'):
            self.headerSubtitle.setText(
                f'<span style="color:{muted};">'
                '엑셀/CSV 파일의 주소 컬럼을 일괄 지오코딩해 포인트 레이어로 추가합니다.'
                '</span>'
            )
        if hasattr(self, 'fileInfoLabel'):
            self.fileInfoLabel.setStyleSheet(f"color: {muted};")
        if hasattr(self, 'autoDetectHint'):
            self.autoDetectHint.setStyleSheet(f"color: {muted};")
        if hasattr(self, 'currentStatusLabel'):
            self.currentStatusLabel.setStyleSheet(f"color: {muted};")

    def _connect_signals(self):
        if hasattr(self, 'mQgsFileWidget'):
            try:
                self.mQgsFileWidget.fileChanged.connect(self._on_file_changed)
            except Exception:
                pass
        if hasattr(self, 'BTNGeoStart'):
            self.BTNGeoStart.clicked.connect(self._on_start_clicked)
        if hasattr(self, 'BTNGeoCancel'):
            self.BTNGeoCancel.clicked.connect(self._on_cancel_clicked)
        if hasattr(self, 'closeBtn'):
            self.closeBtn.clicked.connect(self.accept)
        self.headerCheck.toggled.connect(self._on_header_toggled)

        # 모드 라디오 — 동일 핸들러 호출
        for radio_name in (
            'modeAddress', 'modeCoordXY', 'modeCoordYX', 'modeCoordSeparate',
        ):
            radio = getattr(self, radio_name, None)
            if radio is not None:
                radio.toggled.connect(self._on_mode_changed)

    # ------------------------------------------------------------------
    # 모드 / CRS
    # ------------------------------------------------------------------
    def _populate_crs_combo(self):
        """ KOREA_CRS 상수에서 CRS 콤보 채우기. 기본 선택은 DEFAULT_CRS. """
        if not hasattr(self, 'crsCombo'):
            return
        self.crsCombo.blockSignals(True)
        self.crsCombo.clear()
        default_idx = 0
        for idx, (code, label) in enumerate(KOREA_CRS.items()):
            self.crsCombo.addItem(f"{code} ({label})", code)
            if code == DEFAULT_CRS:
                default_idx = idx
        self.crsCombo.setCurrentIndex(default_idx)
        self.crsCombo.blockSignals(False)

    def _current_mode(self) -> str:
        """ 현재 선택된 입력 모드 상수 반환. """
        if getattr(self, 'modeCoordXY', None) and self.modeCoordXY.isChecked():
            return MODE_XY
        if getattr(self, 'modeCoordYX', None) and self.modeCoordYX.isChecked():
            return MODE_YX
        if getattr(self, 'modeCoordSeparate', None) and self.modeCoordSeparate.isChecked():
            return MODE_SEP
        return MODE_ADDRESS

    def _on_mode_changed(self, _checked: bool = False):
        """ 모드 변경에 따라 컬럼 라벨/Y 슬롯/CRS 활성 상태 갱신. """
        mode = self._current_mode()

        # 컬럼 라벨
        if hasattr(self, 'columnLabel'):
            if mode == MODE_ADDRESS:
                self.columnLabel.setText("주소 컬럼:")
            elif mode == MODE_XY:
                self.columnLabel.setText("좌표 컬럼 (x, y):")
            elif mode == MODE_YX:
                self.columnLabel.setText("좌표 컬럼 (y, x):")
            else:  # MODE_SEP
                self.columnLabel.setText("X 컬럼:")

        # Y 컬럼 슬롯
        show_y = (mode == MODE_SEP)
        for name in ('yColumnLabel', 'yColumnCombo'):
            w = getattr(self, name, None)
            if w is not None:
                w.setVisible(show_y)
                if name == 'yColumnCombo':
                    w.setEnabled(show_y and self._raw_rows != [])

        # CRS 콤보 활성
        if hasattr(self, 'crsCombo'):
            self.crsCombo.setEnabled(mode != MODE_ADDRESS)
        if hasattr(self, 'crsLabel'):
            muted = ThemeColors.muted()
            self.crsLabel.setStyleSheet(
                f"color: {muted};" if mode == MODE_ADDRESS else ""
            )

        # 자동 감지 힌트 텍스트
        if hasattr(self, 'autoDetectHint'):
            if mode == MODE_ADDRESS:
                if self._raw_rows:
                    # 기존 자동 감지 메시지를 _rebuild_view가 다시 세팅
                    self._rebuild_view()
                else:
                    self.autoDetectHint.setText(
                        "파일을 불러오면 가장 가능성 있는 주소 컬럼을 자동 선택합니다."
                    )
            else:
                self.autoDetectHint.setText(
                    "좌표 모드에서는 자동 감지가 적용되지 않습니다. 컬럼을 직접 선택하세요."
                )

    # ------------------------------------------------------------------
    # 파일 선택 / 로드
    # ------------------------------------------------------------------
    def _on_file_changed(self, file_path: str):
        if not file_path or not os.path.exists(file_path):
            self._set_file_info("파일을 선택하면 여기에 행/열 정보가 표시됩니다.")
            return
        try:
            rows = self._read_file(file_path)
            self._raw_rows = rows
            self._column_count = max((len(r) for r in rows), default=0)
            log_info(
                f"파일 로드: {file_path}, 행 {len(rows)}개, 컬럼 {self._column_count}개"
            )
            file_name = os.path.basename(file_path)
            self._set_file_info(
                f"'{file_name}' — {len(rows)}행 × {self._column_count}열을 불러왔습니다."
            )
            self._rebuild_view()
        except Exception as e:
            log_error(f"파일 읽기 실패: {type(e).__name__}: {e}")
            self._set_file_info("파일을 불러올 수 없습니다.")
            self._flash_status(
                "파일을 불러올 수 없습니다. 형식과 인코딩을 확인하세요.",
                level='error',
            )

    def _set_file_info(self, text: str):
        if hasattr(self, 'fileInfoLabel'):
            self.fileInfoLabel.setText(text)

    def _read_file(self, file_path: str) -> List[List[str]]:
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            return self._read_csv(file_path)
        if ext == '.xlsx':
            return self._read_xlsx(file_path)
        if ext == '.xls':
            return self._read_xls(file_path)
        raise ValueError(f"지원하지 않는 파일 형식: {ext}")

    def _read_csv(self, file_path: str) -> List[List[str]]:
        rows: List[List[str]] = []
        last_error: Optional[Exception] = None
        for enc in ('utf-8-sig', 'utf-8', 'cp949', 'euc-kr'):
            try:
                with open(file_path, 'r', encoding=enc, newline='') as f:
                    reader = csv.reader(f)
                    rows = [
                        [(cell if cell is not None else '').strip() for cell in row]
                        for row in reader
                        if any((cell or '').strip() for cell in row)
                    ]
                break
            except UnicodeDecodeError as e:
                last_error = e
                continue
        if not rows and last_error is not None:
            raise last_error
        return rows

    def _read_xlsx(self, file_path: str) -> List[List[str]]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            # openpyxl은 플러그인 libs/ 폴더에 동봉되어 있어 정상 설치라면 항상 로드된다.
            raise ImportError(
                "플러그인에 포함된 엑셀 라이브러리(openpyxl)를 불러오지 못했습니다.\n"
                "플러그인을 삭제한 뒤 다시 설치해 보세요.\n"
                "(그래도 안 되면 OSGeo4W Shell에서 'python -m pip install openpyxl' 실행)"
            )
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            normalized = [
                (str(v).strip() if v is not None else '')
                for v in row
            ]
            if any(normalized):
                rows.append(normalized)
        return rows

    def _read_xls(self, file_path: str) -> List[List[str]]:
        try:
            import xlrd
        except ImportError:
            # xlrd는 플러그인 libs/ 폴더에 동봉되어 있어 정상 설치라면 항상 로드된다.
            raise ImportError(
                "플러그인에 포함된 엑셀 라이브러리(xlrd)를 불러오지 못했습니다.\n"
                "플러그인을 삭제한 뒤 다시 설치해 보세요.\n"
                "(그래도 안 되면 OSGeo4W Shell에서 "
                "'python -m pip install \"xlrd==1.2.0\"' 실행 — "
                "xlrd 2.x는 .xls를 지원하지 않으므로 1.2.0 버전을 명시해야 합니다.)"
            )
        try:
            wb = xlrd.open_workbook(file_path)
        except Exception as e:
            raise ValueError(
                f".xls 파일을 열 수 없습니다: {e}\n"
                "xlrd 2.x가 설치되어 있다면 'pip install \"xlrd==1.2.0\"'으로 다운그레이드하세요."
            )
        ws = wb.sheet_by_index(0)
        rows: List[List[str]] = []
        for r in range(ws.nrows):
            row_vals = []
            for c in range(ws.ncols):
                v = ws.cell_value(r, c)
                if isinstance(v, float) and v.is_integer():
                    v = int(v)
                row_vals.append(str(v).strip() if v != '' else '')
            if any(row_vals):
                rows.append(row_vals)
        return rows

    # ------------------------------------------------------------------
    # 자동 감지 / 미리보기 재구성
    # ------------------------------------------------------------------
    def _score_address_column(self, data_rows: List[List[str]]) -> int:
        """주소 패턴이 가장 많이 매칭되는 컬럼 인덱스 반환"""
        if not data_rows or self._column_count == 0:
            return 0

        best_idx = 0
        best_score = -1
        best_non_empty = -1

        for col in range(self._column_count):
            non_empty = 0
            match_count = 0
            for row in data_rows:
                if col >= len(row):
                    continue
                cell = row[col].strip()
                if not cell:
                    continue
                non_empty += 1
                if cell.replace('.', '', 1).replace('-', '', 1).isdigit():
                    continue
                if ADDR_PATTERN.search(cell):
                    match_count += 1

            score = match_count
            if score > best_score or (
                score == best_score and non_empty > best_non_empty
            ):
                best_score = score
                best_non_empty = non_empty
                best_idx = col

        log_info(
            f"자동 감지: 컬럼 {best_idx} (점수 {best_score}, 비어있지 않음 {best_non_empty})"
        )
        return best_idx

    def _on_header_toggled(self, checked: bool):
        if self._raw_rows:
            self._rebuild_view()

    def _rebuild_view(self):
        """현재 _raw_rows + 헤더 설정으로 미리보기 + 콤보박스 재구성"""
        if not self._raw_rows:
            return

        use_header = self.headerCheck.isChecked()
        N = self._column_count

        if use_header and len(self._raw_rows) > 0:
            first = self._raw_rows[0]
            header_labels = [
                (first[i] if i < len(first) and first[i] else f"{i+1}열")
                for i in range(N)
            ]
            data_rows = self._raw_rows[1:]
        else:
            header_labels = [f"{i+1}열" for i in range(N)]
            data_rows = self._raw_rows

        self._suggested_col = self._score_address_column(data_rows)

        self.columnCombo.blockSignals(True)
        self.columnCombo.clear()
        for i, label in enumerate(header_labels):
            self.columnCombo.addItem(f"{i+1}. {label}", i)
        if 0 <= self._suggested_col < self.columnCombo.count():
            self.columnCombo.setCurrentIndex(self._suggested_col)
        self.columnCombo.setEnabled(True)
        self.columnCombo.setToolTip("지오코딩할 주소(또는 X 좌표)가 들어 있는 컬럼")
        self.columnCombo.blockSignals(False)

        # Y 컬럼 콤보(분리 모드 전용)도 동일 라벨로 채움
        if hasattr(self, 'yColumnCombo'):
            self.yColumnCombo.blockSignals(True)
            self.yColumnCombo.clear()
            for i, label in enumerate(header_labels):
                self.yColumnCombo.addItem(f"{i+1}. {label}", i)
            # 기본값: X 옆 컬럼을 추정 (X+1, 없으면 마지막)
            default_y = min(self._suggested_col + 1, len(header_labels) - 1)
            if default_y < 0:
                default_y = 0
            self.yColumnCombo.setCurrentIndex(default_y)
            self.yColumnCombo.setEnabled(self._current_mode() == MODE_SEP)
            self.yColumnCombo.setToolTip("Y 좌표가 들어 있는 컬럼 (좌표 분리 모드 전용)")
            self.yColumnCombo.blockSignals(False)

        if hasattr(self, 'autoDetectHint'):
            if self._current_mode() == MODE_ADDRESS and 0 <= self._suggested_col < len(header_labels):
                suggested_name = header_labels[self._suggested_col]
                self.autoDetectHint.setText(
                    f"자동 감지: {self._suggested_col + 1}. {suggested_name}  "
                    f"(다른 컬럼이 맞다면 위 콤보박스에서 직접 선택하세요)"
                )
            elif self._current_mode() != MODE_ADDRESS:
                self.autoDetectHint.setText(
                    "좌표 모드에서는 자동 감지가 적용되지 않습니다. 컬럼을 직접 선택하세요."
                )

        all_headers = list(header_labels) + ["상태", "사유", "X", "Y"]
        model = QStandardItemModel(len(data_rows), len(all_headers))
        model.setHorizontalHeaderLabels(all_headers)
        for r_idx, row in enumerate(data_rows):
            for c_idx in range(N):
                value = row[c_idx] if c_idx < len(row) else ""
                model.setItem(r_idx, c_idx, QStandardItem(value))
            model.setItem(r_idx, N, QStandardItem("대기"))
            model.setItem(r_idx, N + 1, QStandardItem(""))
            model.setItem(r_idx, N + 2, QStandardItem(""))
            model.setItem(r_idx, N + 3, QStandardItem(""))
        self.tableView.setModel(model)
        self._preview_model = model

    # ------------------------------------------------------------------
    # 지오코딩 실행
    # ------------------------------------------------------------------
    def _on_start_clicked(self, checked=False):
        if not self._raw_rows:
            self._flash_status("먼저 엑셀/CSV 파일을 선택해주세요.", level='warn')
            return
        if self.worker is not None and self.worker.isRunning():
            self._flash_status("이미 지오코딩이 진행 중입니다.", level='warn')
            return

        mode = self._current_mode()
        if mode == MODE_ADDRESS:
            if not self._require_user_api_key():
                return
            self._run_address_mode()
        else:
            self._run_coordinate_mode(mode)

    def _require_user_api_key(self) -> bool:
        """ 사용자 API 키 검증 — 주소 모드 진입 직전에 호출. """
        try:
            key = ConfigManager().user_api_key
        except Exception:
            key = ''
        if not key:
            self.show_warning_message(
                "사용자 API 키 필요",
                ERROR_MESSAGES['user_api_key_missing'],
            )
            return False
        return True

    def _run_address_mode(self):
        """ 기존 주소 모드 — V-World API + GeocodingWorker. """
        col_idx = self.columnCombo.currentIndex()
        if col_idx < 0:
            self._flash_status("주소 컬럼을 선택해주세요.", level='warn')
            return

        use_header = self.headerCheck.isChecked()
        data_rows = self._raw_rows[1:] if use_header else self._raw_rows

        addresses: List[str] = []
        self._selected_data_row_indices: List[int] = []
        for d_idx, row in enumerate(data_rows):
            if col_idx >= len(row):
                continue
            value = row[col_idx].strip()
            if not value:
                continue
            addresses.append(value)
            self._selected_data_row_indices.append(d_idx)

        if not addresses:
            self._flash_status(
                "선택한 컬럼에 지오코딩할 주소가 없습니다. 다른 컬럼을 선택해보세요.",
                level='warn',
            )
            return

        log_info(
            f"지오코딩 시작(주소): 컬럼 {col_idx}, 헤더={use_header}, {len(addresses)}건"
        )

        self.worker = GeocodingWorker(addresses, DEFAULT_CRS)
        self.worker.progress.connect(self.geocoderProgressBar.setValue)
        self.worker.status.connect(self._on_worker_status)
        self.worker.finished.connect(self._on_finished)
        self.worker.error.connect(self._on_error)

        self.geocoderProgressBar.setValue(0)
        self._set_running(True)
        self._flash_status(f"{len(addresses)}개 주소 지오코딩을 시작합니다.", level='info')
        self.worker.start()

    def _run_coordinate_mode(self, mode: str):
        """ 좌표 모드 — API 없이 직접 파싱하여 포인트 레이어 생성(동기). """
        col_x = self.columnCombo.currentIndex()
        if col_x < 0:
            self._flash_status("좌표 컬럼을 선택해주세요.", level='warn')
            return

        col_y = None
        if mode == MODE_SEP:
            col_y = self.yColumnCombo.currentIndex()
            if col_y < 0:
                self._flash_status("Y 컬럼을 선택해주세요.", level='warn')
                return
            if col_x == col_y:
                self._flash_status("X 컬럼과 Y 컬럼이 같습니다. 다른 컬럼을 선택하세요.", level='warn')
                return

        crs = (self.crsCombo.currentData() if hasattr(self, 'crsCombo') else None) or DEFAULT_CRS

        use_header = self.headerCheck.isChecked()
        data_rows = self._raw_rows[1:] if use_header else self._raw_rows

        results: list = []
        self._selected_data_row_indices = []
        success_n = 0
        for d_idx, row in enumerate(data_rows):
            x, y, err = self._extract_coords(row, mode, col_x, col_y)
            summary = self._row_summary(row, mode, col_x, col_y)
            if err:
                results.append({'address': summary, 'x': 0, 'y': 0, 'status': err})
            else:
                results.append({'address': summary, 'x': x, 'y': y, 'status': '성공'})
                success_n += 1
            self._selected_data_row_indices.append(d_idx)

        if not results:
            self._flash_status("처리할 데이터가 없습니다.", level='warn')
            return

        if success_n == 0:
            self._flash_status(
                "모든 행에서 좌표 파싱에 실패했습니다. 컬럼/모드/구분자를 확인하세요.",
                level='warn',
            )

        log_info(
            f"지오코딩 시작(좌표 {mode}): {len(results)}건 처리, 성공 {success_n}, CRS={crs}"
        )

        self.geocoderProgressBar.setValue(100)
        # 기존 finished 핸들러 재사용 (crs 인자로 레이어 좌표계 변경)
        self._on_finished(results, crs=crs)

    def _extract_coords(self, row, mode, col_x, col_y):
        """ 행에서 (x, y) 추출. 실패 시 (None, None, 사유) 반환. """
        try:
            if mode == MODE_SEP:
                if col_x >= len(row) or col_y >= len(row):
                    return None, None, "셀 범위 밖"
                x_raw = (row[col_x] or '').strip()
                y_raw = (row[col_y] or '').strip()
                if not x_raw or not y_raw:
                    return None, None, "좌표 셀이 비어 있음"
                return float(x_raw), float(y_raw), None
            # 단일 셀 모드
            if col_x >= len(row):
                return None, None, "셀 범위 밖"
            cell = (row[col_x] or '').strip()
            if not cell:
                return None, None, "좌표 셀이 비어 있음"
            tokens = [t for t in COORD_SPLIT_RE.split(cell) if t]
            if len(tokens) < 2:
                return None, None, f"좌표 토큰 부족: '{cell}'"
            a, b = float(tokens[0]), float(tokens[1])
            if mode == MODE_YX:
                return b, a, None  # y,x → x,y
            return a, b, None      # x,y
        except ValueError as e:
            return None, None, f"잘못된 좌표 형식: {e}"
        except Exception as e:
            return None, None, f"파싱 오류: {type(e).__name__}: {e}"

    @staticmethod
    def _row_summary(row, mode, col_x, col_y) -> str:
        """ 결과 레이어의 '주소' 필드에 들어갈 행 요약 문자열. """
        try:
            if mode == MODE_SEP and col_y is not None:
                xv = row[col_x] if col_x < len(row) else ''
                yv = row[col_y] if col_y < len(row) else ''
                return f"X={xv}, Y={yv}"
            return (row[col_x] if col_x < len(row) else '') or ''
        except Exception:
            return ''

    def _on_cancel_clicked(self, checked=False):
        if self.worker and self.worker.isRunning():
            self.worker.cancel()
            self._flash_status("작업을 취소했습니다.", level='warn')
            self._set_running(False)

    def _set_running(self, running: bool):
        """시작/취소 버튼 상태 + 진행 라벨 토글"""
        if hasattr(self, 'BTNGeoStart'):
            self.BTNGeoStart.setEnabled(not running)
        if hasattr(self, 'BTNGeoCancel'):
            self.BTNGeoCancel.setVisible(running)
        if not running and hasattr(self, 'currentStatusLabel'):
            self.currentStatusLabel.clear()

    def _on_worker_status(self, text: str):
        if hasattr(self, 'currentStatusLabel'):
            self.currentStatusLabel.setText(text)

    def _on_progress(self, value: int):
        self.geocoderProgressBar.setValue(value)

    def _on_error(self, msg: str):
        self._set_running(False)
        log_error(f"지오코딩 오류: {msg}")
        self._flash_status(f"지오코딩 오류: {msg}", level='error')
        self.show_error_message("지오코딩 오류", msg)

    def _on_finished(self, results: list, crs: str = DEFAULT_CRS):
        self._set_running(False)
        log_info(f"지오코딩 완료: {len(results)}개 결과, CRS={crs}")

        status_col = self._column_count
        reason_col = self._column_count + 1
        x_col = self._column_count + 2
        y_col = self._column_count + 3

        success_count = 0
        if self._preview_model is not None:
            indices = getattr(self, '_selected_data_row_indices', list(range(len(results))))
            for r_idx, result in zip(indices, results):
                if r_idx >= self._preview_model.rowCount():
                    continue
                raw_status = str(result.get('status', ''))
                is_success = raw_status == '성공'
                display_status = "성공" if is_success else "실패"
                reason = "" if is_success else raw_status
                x_text = str(result.get('x', '')) if is_success else ""
                y_text = str(result.get('y', '')) if is_success else ""
                if is_success:
                    success_count += 1

                self._preview_model.setItem(r_idx, status_col, QStandardItem(display_status))
                self._preview_model.setItem(r_idx, reason_col, QStandardItem(reason))
                self._preview_model.setItem(r_idx, x_col, QStandardItem(x_text))
                self._preview_model.setItem(r_idx, y_col, QStandardItem(y_text))

        try:
            success = self._create_layer(results, crs=crs)
            self._flash_status(
                f"완료 — 성공 {success} / 전체 {len(results)}",
                level='info',
            )
            self.show_success_message(
                "완료",
                f"지오코딩 완료\n성공: {success} / 전체: {len(results)}\n"
                f"'{GEOCODER_LAYER}' 레이어가 추가되었습니다."
            )
        except Exception as e:
            log_error(f"레이어 생성 실패: {type(e).__name__}: {e}")
            self._flash_status(
                "레이어 생성에 실패했습니다. QGIS 로그를 확인하세요.",
                level='error',
            )
            self.show_error_message("레이어 생성 실패", str(e))

    def _create_layer(self, results: list, crs: str = DEFAULT_CRS) -> int:
        layer = QgsVectorLayer(f"Point?crs={crs}", GEOCODER_LAYER, "memory")
        provider = layer.dataProvider()

        provider.addAttributes([
            QgsField("주소", _vtype('String')),
            QgsField("상태", _vtype('String')),
            QgsField("사유", _vtype('String')),
            QgsField("X", _vtype('Double')),
            QgsField("Y", _vtype('Double')),
        ])
        layer.updateFields()

        features = []
        success = 0
        for r in results:
            raw_status = str(r.get('status', ''))
            is_success = raw_status == '성공'

            x_attr = None
            y_attr = None
            geometry = None

            if is_success:
                try:
                    x_val = float(r['x'])
                    y_val = float(r['y'])
                    geometry = QgsGeometry.fromPointXY(QgsPointXY(x_val, y_val))
                    x_attr = x_val
                    y_attr = y_val
                    success += 1
                except (KeyError, ValueError, TypeError) as e:
                    is_success = False
                    raw_status = f"좌표 변환 실패: {e}"

            feat = QgsFeature(layer.fields())
            if geometry is not None:
                feat.setGeometry(geometry)
            feat.setAttribute("주소", r.get('address', ''))
            feat.setAttribute("상태", "성공" if is_success else "실패")
            feat.setAttribute("사유", "" if is_success else raw_status)
            feat.setAttribute("X", x_attr)
            feat.setAttribute("Y", y_attr)
            features.append(feat)

        if features:
            provider.addFeatures(features)
            layer.updateExtents()

        QgsProject.instance().addMapLayer(layer)
        return success

    # ------------------------------------------------------------------
    # 인라인 상태 표시
    # ------------------------------------------------------------------
    def _flash_status(self, message: str, level: str = 'info'):
        if not hasattr(self, 'statusLabel'):
            return
        self.statusLabel.setStyleSheet(f"color: {_status_color(level)};")
        self.statusLabel.setText(message)
        # 오류는 사용자가 읽을 시간을 확보하도록 자동으로 사라지지 않게 한다.
        if level == 'error':
            self._status_timer.stop()
        else:
            self._status_timer.start(_STATUS_FADE_MS)

    def _clear_status(self):
        if hasattr(self, 'statusLabel'):
            self.statusLabel.clear()
